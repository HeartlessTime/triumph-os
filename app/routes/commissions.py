import os
import shutil
from datetime import datetime
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, Form, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.commission_entry import CommissionEntry
from app.template_config import templates

router = APIRouter(prefix="/commissions", tags=["commissions"])

UPLOAD_DIR = os.getenv("UPLOAD_DIR", "./uploads")
TEMPLATE_FILENAME = "commission_template.xlsx"


def _template_path() -> str:
    return os.path.join(UPLOAD_DIR, TEMPLATE_FILENAME)


def _template_exists() -> bool:
    return os.path.isfile(_template_path())


def _parse_decimal(value: str | None) -> Decimal | None:
    if not value or value.strip() == "":
        return None
    try:
        return Decimal(value.strip().replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


# ---------------------------------------------------------------------------
# List view
# ---------------------------------------------------------------------------
@router.get("", response_class=HTMLResponse)
async def commission_list(request: Request, db: Session = Depends(get_db)):
    month = request.query_params.get("month")
    if not month:
        month = datetime.utcnow().strftime("%Y-%m")

    entries = (
        db.query(CommissionEntry)
        .filter(CommissionEntry.month == month)
        .order_by(CommissionEntry.account_name, CommissionEntry.job_name)
        .all()
    )

    return templates.TemplateResponse(
        "commissions/list.html",
        {
            "request": request,
            "entries": entries,
            "selected_month": month,
            "has_template": _template_exists(),
        },
    )


# ---------------------------------------------------------------------------
# Add entry
# ---------------------------------------------------------------------------
@router.post("/add")
async def add_commission(
    request: Request,
    month: str = Form(...),
    account_name: str = Form(...),
    job_name: str = Form(...),
    job_number: str = Form(""),
    contact: str = Form(""),
    job_amount: str = Form(""),
    commission_amount: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    entry = CommissionEntry(
        month=month.strip(),
        account_name=account_name.strip(),
        job_name=job_name.strip(),
        job_number=job_number.strip() or None,
        contact=contact.strip() or None,
        job_amount=_parse_decimal(job_amount),
        commission_amount=_parse_decimal(commission_amount),
        notes=notes.strip() or None,
    )
    db.add(entry)
    db.commit()
    return RedirectResponse(url=f"/commissions?month={entry.month}", status_code=303)


# ---------------------------------------------------------------------------
# Edit entry
# ---------------------------------------------------------------------------
@router.post("/{entry_id}/edit")
async def edit_commission(
    entry_id: int,
    request: Request,
    month: str = Form(...),
    account_name: str = Form(...),
    job_name: str = Form(...),
    job_number: str = Form(""),
    contact: str = Form(""),
    job_amount: str = Form(""),
    commission_amount: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    entry = db.query(CommissionEntry).filter(CommissionEntry.id == entry_id).first()
    if not entry:
        return RedirectResponse(url="/commissions", status_code=303)

    entry.month = month.strip()
    entry.account_name = account_name.strip()
    entry.job_name = job_name.strip()
    entry.job_number = job_number.strip() or None
    entry.contact = contact.strip() or None
    entry.job_amount = _parse_decimal(job_amount)
    entry.commission_amount = _parse_decimal(commission_amount)
    entry.notes = notes.strip() or None
    entry.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(url=f"/commissions?month={entry.month}", status_code=303)


# ---------------------------------------------------------------------------
# Delete entry
# ---------------------------------------------------------------------------
@router.post("/{entry_id}/delete")
async def delete_commission(
    entry_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    entry = db.query(CommissionEntry).filter(CommissionEntry.id == entry_id).first()
    redirect_month = entry.month if entry else datetime.utcnow().strftime("%Y-%m")
    if entry:
        db.delete(entry)
        db.commit()
    return RedirectResponse(url=f"/commissions?month={redirect_month}", status_code=303)


# ---------------------------------------------------------------------------
# Upload Excel template
# ---------------------------------------------------------------------------
@router.post("/upload-template")
async def upload_template(
    request: Request,
    file: UploadFile = File(...),
):
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    dest = _template_path()
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    from_month = request.query_params.get("month", datetime.utcnow().strftime("%Y-%m"))
    return RedirectResponse(url=f"/commissions?month={from_month}", status_code=303)


# ---------------------------------------------------------------------------
# Export to Excel
# ---------------------------------------------------------------------------
@router.post("/export")
async def export_commissions(
    request: Request,
    export_month: str = Form(...),
    db: Session = Depends(get_db),
):
    if not _template_exists():
        return RedirectResponse(url=f"/commissions?month={export_month}", status_code=303)

    entries = (
        db.query(CommissionEntry)
        .filter(CommissionEntry.month == export_month)
        .order_by(CommissionEntry.account_name, CommissionEntry.job_name)
        .all()
    )

    if not entries:
        return RedirectResponse(url=f"/commissions?month={export_month}", status_code=303)

    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(_template_path())
    ws = wb.active

    # ---- Row 5: Update month label ----
    # Parse "YYYY-MM" into a readable label like "February 2026"
    try:
        dt = datetime.strptime(export_month, "%Y-%m")
        month_label = dt.strftime("%B %Y")
    except ValueError:
        month_label = export_month

    # Write only the value to row 5, column A — preserves formatting
    ws.cell(row=5, column=1).value = month_label

    # ---- Find first empty row starting at row 11 ----
    start_row = 11
    while True:
        row_empty = True
        for col in range(1, 8):  # Columns A–G
            if ws.cell(row=start_row, column=col).value is not None:
                row_empty = False
                break
        if row_empty:
            break
        start_row += 1

    # ---- Write commission data ----
    # Fixed column mapping:
    #   A = account_name
    #   B = job_name
    #   C = job_number
    #   D = contact
    #   E = job_amount
    #   F = commission_amount
    #   G = notes
    for i, entry in enumerate(entries):
        row = start_row + i
        ws.cell(row=row, column=1).value = entry.account_name
        ws.cell(row=row, column=2).value = entry.job_name
        ws.cell(row=row, column=3).value = entry.job_number
        ws.cell(row=row, column=4).value = entry.contact
        ws.cell(row=row, column=5).value = float(entry.job_amount) if entry.job_amount is not None else None
        ws.cell(row=row, column=6).value = float(entry.commission_amount) if entry.commission_amount is not None else None
        ws.cell(row=row, column=7).value = entry.notes

    # ---- Save to BytesIO ----
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # ---- Mark entries as exported ----
    for entry in entries:
        entry.status = "exported"
        entry.exported_month = export_month
    db.commit()

    # ---- Build filename ----
    filename = f"{month_label.replace(' ', '_')}_Commission_Sheet.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
